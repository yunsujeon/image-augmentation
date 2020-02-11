import cv2
import numpy as np
import os
from openpyxl import load_workbook
from openpyxl import Workbook

try:
    import Image
except ImportError:
    from PIL import Image

# # 엑셀 불러오기
# load_wb = load_workbook(filename='C:/Users/jeon/Desktop/pad.xlsx')
# Sheet1 = load_wb['Sheet']
# # 셀 좌표로 값 출력
# a = Sheet1.cell(1, 2).value
# print(a)
#
# # 엑셀에 쓰기
# write_wb = Workbook()
# # Sheet1에다 입력
# Sheet1 = write_wb.active
# # 셀 단위로 추가
# row_index = 1
# for i in range(1, 10):
#     Sheet1.cell(i, i, '{:d}'.format(i))
#     Sheet1.cell(row=row_index, column=1).value = i
#     row_index = row_index + 1
# write_wb.save(filename='C:/Users/jeon/Desktop/pad.xlsx') # 열려으면 써지지 않음.


path_dir = 'C:/Users/jeon/Desktop/aaa/'
file_list = os.listdir(path_dir)  # path에 존재하는 파일이름들 불러오기
file_list.sort()  # 파일 이름대로 정렬

for item in file_list:  # file_list = ['a','b', ~~~]
    if item.find('bmp') is not -1:  # A-54포함하는 이미지만 load
        if item.find('RGB') is not -1:
            rgbname = os.path.splitext(item)[0]
            rgbimage = cv2.imread((path_dir + item), cv2.IMREAD_COLOR)
            rgbreplace = rgbname.replace("_RGB", "_")
    if item.find('bmp') is not -1:
        if item.find('SolderMask') is not -1:  # 그중에 soldermask만 걸러낸다.
            mskname = os.path.splitext(item)[0]
            mskimage = cv2.imread((path_dir + item), cv2.IMREAD_GRAYSCALE)
            mskreplace = mskname.replace("_SolderMask", "_")

            # 화소별로 접근하여 left right top bottom 과 이의 center를 찾아낸다.
            height, width = mskimage.shape

            # 각 요소들 검출 T/F find
            tfind = True
            bfind = True
            lfind = True
            rfind = True

            for y in range(0, height - 1):  # top 검출
                if tfind is False:
                    break
                else:
                    for x in range(0, width - 1):
                        pixel = mskimage.item(y, x)
                        if pixel is 255:
                            t = y
                            tfind = False
                            break

            for y in range(height - 1, 0, -1):  # b 검출
                if bfind is False:
                    break
                else:
                    for x in range(width - 1, 0, -1):
                        pixel = mskimage.item(y, x)
                        if pixel is 255:
                            b = y
                            bfind = False
                            break

            for x in range(0, width - 1):  # l 검출
                if lfind is False:
                    break
                else:
                    for y in range(0, height - 1):
                        pixel = mskimage.item(y, x)
                        if pixel is 255:
                            l = x
                            lfind = False
                            break

            for x in range(width - 1, 0, -1):  # r 검출
                if rfind is False:
                    break
                else:
                    for y in range(height - 1, 0, -1):
                        pixel = mskimage.item(y, x)
                        if pixel is 255:
                            r = x
                            rfind = False
                            break

            b = b + 1  # crop시 한픽셀씩 덜 저장하기때문에 미리 +1
            r = r + 1
            count = 0  # 저장되는 filename을 서로 다르게 해주기 위함
            number = 5  # 한번 움직일 때마다 픽셀 이동 수
            for iter_aug in range(1, 6):

                tmn = t - iter_aug * number  # tmn = t - number
                bmn = b - iter_aug * number  # bmn = b - number
                tpn = t + iter_aug * number  # tpn = t + number
                bpn = b + iter_aug * number  # bpn = b + number
                lmn = l - iter_aug * number  # lmn = l - number
                rmn = r - iter_aug * number  #
                lpn = l + iter_aug * number  #
                rpn = r + iter_aug * number  #

                if tmn >= 0 and bmn > t:  # image 크기 넘어가지 않는 조건 / 빈 셀 crop 하지 않는 조건
                    mskimg1 = mskimage[tmn: bmn, l: r]  # 위로 number픽셀씩 이동
                    rgbimg1 = rgbimage[tmn: bmn, l: r]
                    mskpath = (path_dir + mskreplace + '{:d}'.format(count) + '_SolderMask.bmp')  # 단순 경로 지정
                    rgbpath = (path_dir + rgbreplace + '{:d}'.format(count) + '_RGB.bmp')
                    cv2.imwrite(mskpath, mskimg1)  # crop 한 이미지 저장
                    cv2.imwrite(rgbpath, rgbimg1)
                    count = count + 1
                if bpn <= height and tpn < b:
                    mskimg2 = mskimage[tpn: bpn, l: r]  # 아래로 5픽셀씩 이동
                    rgbimg2 = rgbimage[tpn: bpn, l:r]
                    mskpath = (path_dir + mskreplace + '{:d}'.format(count) + '_SolderMask.bmp')
                    rgbpath = (path_dir + rgbreplace + '{:d}'.format(count) + '_RGB.bmp')
                    cv2.imwrite(mskpath, mskimg2)
                    cv2.imwrite(rgbpath, rgbimg2)
                    count = count + 1
                if lmn >= 0 and rmn > l:
                    mskimg3 = mskimage[t: b, lmn: rmn]  # 왼쪽으로 5픽셀씩 이동
                    rgbimg3 = rgbimage[t: b, lmn: rmn]
                    mskpath = (path_dir + mskreplace + '{:d}'.format(count) + '_SolderMask.bmp')
                    rgbpath = (path_dir + rgbreplace + '{:d}'.format(count) + '_RGB.bmp')
                    cv2.imwrite(mskpath, mskimg3)
                    cv2.imwrite(rgbpath, rgbimg3)
                    count = count + 1
                if rpn <= width and lpn < r:
                    mskimg4 = mskimage[t: b, lpn: rpn]  # 오른쪽으로 5픽셀씩 이동
                    rgbimg4 = rgbimage[t: b, lpn: rpn]
                    mskpath = (path_dir + mskreplace + '{:d}'.format(count) + '_SolderMask.bmp')
                    rgbpath = (path_dir + rgbreplace + '{:d}'.format(count) + '_RGB.bmp')
                    cv2.imwrite(mskpath, mskimg4)
                    cv2.imwrite(rgbpath, rgbimg4)
                    count = count + 1
